"""Turns one sustained RoutingComparison run into charts, a screenshot and a workbook.

Input is benchmark-run.txt, the captured output of a 90 seconds per stage run against a
FalkorDB Cloud instance with one primary and one replica.

Every figure here is measured. Nothing is modelled or extrapolated.
"""

import csv
import io
import re
from pathlib import Path

import matplotlib

matplotlib.use("Agg")
import matplotlib.pyplot as plt
from matplotlib.ticker import FuncFormatter
from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill

HERE = Path(__file__).resolve().parent
SOURCE = HERE / "benchmark-run.txt"

PRIMARY_COLOUR = "#d1495b"
ROTATE_COLOUR = "#00916e"
REPLICA_COLOUR = "#5aa9e6"
IDLE_COLOUR = "#d9dcd6"

FALKOR_THREAD_COUNT = 2

MODES = ("PRIMARY_ONLY", "ROUND_ROBIN")
MODE_LABELS = {
    "PRIMARY_ONLY": "Primary only",
    "ROUND_ROBIN": "Rotating primary and replica",
}


def read_rows(path):
    """Pulls the CSV block out of the captured console output."""
    text = path.read_text()
    marker = text.index("===== CSV =====")
    block = text[marker:].split("\n", 1)[1]
    lines = []
    for line in block.splitlines():
        line = line.strip()
        if not line:
            break
        lines.append(line)
    reader = csv.DictReader(io.StringIO("\n".join(lines)))
    rows = {}
    for row in reader:
        key = (int(row["threads"]), row["mode"])
        rows[key] = {
            "reads": int(row["reads"]),
            "seconds": float(row["seconds"]),
            "throughput": float(row["throughput_qps"]),
            "latency": float(row["latency_ms"]),
            "primary_cores": float(row["primary_cores"]),
            "replica_cores": float(row["replica_cores"]),
            "failures": int(row["failures"]),
        }
    return rows


def read_context(path):
    """Grabs the dataset size line so the artifacts state what was queried."""
    match = re.search(r"about ([\d,]+) nodes total", path.read_text())
    return match.group(1) if match else "unknown"


def thread_counts(rows):
    return sorted({threads for threads, _ in rows})


def thousands(value, _pos):
    return f"{value:,.0f}"


def make_charts(rows, nodes, out):
    threads = thread_counts(rows)
    positions = range(len(threads))
    width = 0.36

    figure, axes = plt.subplots(2, 2, figsize=(15, 10))
    figure.suptitle(
        "Reading from the replica as well as the primary\n"
        f"FalkorDB Cloud, AWS us-east-2, 1 primary + 1 replica, 2 cores per DB pod, "
        f"4 core client\n{nodes} nodes, 90 seconds sustained load per bar",
        fontsize=15,
        fontweight="bold",
    )

    throughput = axes[0][0]
    for offset, mode, colour in (
        (-width / 2, "PRIMARY_ONLY", PRIMARY_COLOUR),
        (width / 2, "ROUND_ROBIN", ROTATE_COLOUR),
    ):
        values = [rows[(count, mode)]["throughput"] for count in threads]
        bars = throughput.bar(
            [position + offset for position in positions],
            values,
            width,
            label=MODE_LABELS[mode],
            color=colour,
        )
        throughput.bar_label(bars, fmt="{:,.0f}", fontsize=9, padding=2)
    throughput.set_title("Throughput", fontweight="bold")
    throughput.set_ylabel("reads per second")
    throughput.set_xlabel("client threads")
    throughput.set_xticks(list(positions))
    throughput.set_xticklabels(threads)
    throughput.yaxis.set_major_formatter(FuncFormatter(thousands))
    throughput.legend(loc="upper left")
    throughput.grid(axis="y", alpha=0.3)
    for index, count in enumerate(threads):
        primary = rows[(count, "PRIMARY_ONLY")]["throughput"]
        rotate = rows[(count, "ROUND_ROBIN")]["throughput"]
        gain = (rotate / primary - 1) * 100
        throughput.annotate(
            f"{gain:+.0f}%",
            (index, max(primary, rotate)),
            textcoords="offset points",
            xytext=(0, 26),
            ha="center",
            fontsize=12,
            fontweight="bold",
            color=ROTATE_COLOUR if gain > 5 else "#6c757d",
        )
    throughput.set_ylim(0, max(row["throughput"] for row in rows.values()) * 1.28)

    utilisation = axes[0][1]
    labels = []
    primary_cores = []
    replica_cores = []
    for count in threads:
        for mode in MODES:
            row = rows[(count, mode)]
            labels.append(f"{count}t\n{'primary' if mode == 'PRIMARY_ONLY' else 'rotate'}")
            primary_cores.append(row["primary_cores"])
            replica_cores.append(row["replica_cores"])
    spots = range(len(labels))
    utilisation.bar(
        [spot - width / 2 for spot in spots],
        primary_cores,
        width,
        label="primary node",
        color=PRIMARY_COLOUR,
    )
    utilisation.bar(
        [spot + width / 2 for spot in spots],
        replica_cores,
        width,
        label="replica node",
        color=REPLICA_COLOUR,
    )
    utilisation.axhline(
        FALKOR_THREAD_COUNT,
        color="#333333",
        linestyle="--",
        linewidth=1.2,
    )
    utilisation.annotate(
        f"THREAD_COUNT={FALKOR_THREAD_COUNT}, one node saturated",
        (len(labels) - 0.5, FALKOR_THREAD_COUNT),
        textcoords="offset points",
        xytext=(-6, 6),
        ha="right",
        fontsize=9,
    )
    utilisation.set_title(
        "Utilization: cores busy per node (this is the point)", fontweight="bold"
    )
    utilisation.set_ylabel("cores busy")
    utilisation.set_xticks(list(spots))
    utilisation.set_xticklabels(labels, fontsize=9)
    utilisation.set_ylim(0, FALKOR_THREAD_COUNT * 1.25)
    utilisation.legend(loc="upper left")
    utilisation.grid(axis="y", alpha=0.3)

    latency = axes[1][0]
    for mode, colour in ((("PRIMARY_ONLY"), PRIMARY_COLOUR), ("ROUND_ROBIN", ROTATE_COLOUR)):
        values = [rows[(count, mode)]["latency"] for count in threads]
        latency.plot(
            threads,
            values,
            marker="o",
            linewidth=2.2,
            label=MODE_LABELS[mode],
            color=colour,
        )
        for count, value in zip(threads, values):
            latency.annotate(
                f"{value:.2f}",
                (count, value),
                textcoords="offset points",
                xytext=(0, 9),
                ha="center",
                fontsize=9,
            )
    latency.set_title("Mean client latency", fontweight="bold")
    latency.set_ylabel("milliseconds")
    latency.set_xlabel("client threads")
    latency.set_xscale("log", base=2)
    latency.set_xticks(threads)
    latency.set_xticklabels(threads)
    latency.legend()
    latency.grid(alpha=0.3)

    rejected = axes[1][1]
    for offset, mode, colour in (
        (-width / 2, "PRIMARY_ONLY", PRIMARY_COLOUR),
        (width / 2, "ROUND_ROBIN", ROTATE_COLOUR),
    ):
        values = [rows[(count, mode)]["failures"] for count in threads]
        bars = rejected.bar(
            [position + offset for position in positions],
            values,
            width,
            label=MODE_LABELS[mode],
            color=colour,
        )
        rejected.bar_label(bars, fmt="{:,.0f}", fontsize=9, padding=2)
    rejected.set_title(
        "Queries rejected with 'Max pending queries exceeded'", fontweight="bold"
    )
    rejected.set_ylabel("rejected queries")
    rejected.set_xlabel("client threads")
    rejected.set_xticks(list(positions))
    rejected.set_xticklabels(threads)
    rejected.yaxis.set_major_formatter(FuncFormatter(thousands))
    rejected.legend()
    rejected.grid(axis="y", alpha=0.3)
    rejected.set_ylim(0, max(row["failures"] for row in rows.values()) * 1.25 or 1)

    figure.tight_layout(rect=(0, 0, 1, 0.93))
    figure.savefig(out, dpi=160)
    plt.close(figure)


def make_terminal_screenshot(source, out):
    """Renders the console output as a terminal window image."""
    keep = []
    for line in source.read_text().splitlines():
        if line.startswith("  seeded ") or line.startswith("master '"):
            continue
        keep.append(line)
    while keep and not keep[0].strip():
        keep.pop(0)
    body = "\n".join(keep)

    height = 0.185 * (len(keep) + 3)
    figure = plt.figure(figsize=(12.5, height), facecolor="#1d1f21")
    figure.text(
        0.012,
        0.995,
        "ubuntu@ip-172-31-24-11:~/replica-read-routing$ mvn exec:java "
        "-Dexec.mainClass=...RoutingComparison",
        family="monospace",
        fontsize=9.5,
        color="#8abeb7",
        va="top",
    )
    figure.text(
        0.012,
        0.955,
        body,
        family="monospace",
        fontsize=9.5,
        color="#c5c8c6",
        va="top",
        linespacing=1.42,
    )
    figure.savefig(out, dpi=160, facecolor="#1d1f21")
    plt.close(figure)


def make_workbook(rows, nodes, out):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Results"

    title = Font(bold=True, size=14)
    header = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F3A5F")
    measured_fill = PatternFill("solid", fgColor="E8F5E9")

    sheet["A1"] = "Replica read routing, sustained load"
    sheet["A1"].font = title
    sheet["A2"] = (
        f"FalkorDB Cloud, AWS us-east-2, 1 primary + 1 replica, 2 cores per DB pod, "
        f"THREAD_COUNT={FALKOR_THREAD_COUNT} per node, 4 core client (EC2 c4.xlarge), "
        f"{nodes} nodes, 90 seconds per row."
    )
    sheet["A3"] = "Every cell is measured. Nothing on this sheet is modelled or extrapolated."
    sheet["A3"].font = Font(italic=True, color="2E7D32")

    columns = [
        ("Threads", "threads"),
        ("Mode", "mode"),
        ("Reads", "reads"),
        ("Reads/sec", "throughput"),
        ("Mean latency ms", "latency"),
        ("Primary cores busy", "primary_cores"),
        ("Replica cores busy", "replica_cores"),
        ("Rejected queries", "failures"),
    ]
    start = 5
    for index, (label, _) in enumerate(columns, start=1):
        cell = sheet.cell(row=start, column=index, value=label)
        cell.font = header
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    row_number = start + 1
    for count in thread_counts(rows):
        for mode in MODES:
            data = rows[(count, mode)]
            values = [
                count,
                MODE_LABELS[mode],
                data["reads"],
                data["throughput"],
                data["latency"],
                data["primary_cores"],
                data["replica_cores"],
                data["failures"],
            ]
            for index, value in enumerate(values, start=1):
                cell = sheet.cell(row=row_number, column=index, value=value)
                cell.fill = measured_fill
                if index in (3, 4, 8):
                    cell.number_format = "#,##0"
                if index in (5, 6, 7):
                    cell.number_format = "0.00"
            row_number += 1

    gain_row = row_number + 1
    sheet.cell(row=gain_row, column=1, value="Throughput gain from rotating").font = Font(bold=True)
    gain_row += 1
    sheet.cell(row=gain_row, column=1, value="Threads").font = header
    sheet.cell(row=gain_row, column=1).fill = header_fill
    sheet.cell(row=gain_row, column=2, value="Gain").font = header
    sheet.cell(row=gain_row, column=2).fill = header_fill
    for count in thread_counts(rows):
        gain_row += 1
        primary = rows[(count, "PRIMARY_ONLY")]["throughput"]
        rotate = rows[(count, "ROUND_ROBIN")]["throughput"]
        sheet.cell(row=gain_row, column=1, value=count)
        cell = sheet.cell(row=gain_row, column=2, value=rotate / primary - 1)
        cell.number_format = "+0.0%;-0.0%"

    for column, width in zip("ABCDEFGH", (9, 30, 14, 12, 16, 18, 18, 16)):
        sheet.column_dimensions[column].width = width

    last = start + len(rows)

    throughput_chart = BarChart()
    throughput_chart.type = "col"
    throughput_chart.title = "Reads per second"
    throughput_chart.y_axis.title = "reads/sec"
    throughput_chart.add_data(
        Reference(sheet, min_col=4, min_row=start, max_row=last), titles_from_data=True
    )
    throughput_chart.set_categories(
        Reference(sheet, min_col=1, max_col=2, min_row=start + 1, max_row=last)
    )
    throughput_chart.height = 9
    throughput_chart.width = 20
    sheet.add_chart(throughput_chart, "J5")

    cores_chart = BarChart()
    cores_chart.type = "col"
    cores_chart.title = "Cores busy per node"
    cores_chart.y_axis.title = "cores busy"
    cores_chart.add_data(
        Reference(sheet, min_col=6, max_col=7, min_row=start, max_row=last),
        titles_from_data=True,
    )
    cores_chart.set_categories(
        Reference(sheet, min_col=1, max_col=2, min_row=start + 1, max_row=last)
    )
    cores_chart.height = 9
    cores_chart.width = 20
    sheet.add_chart(cores_chart, "J24")

    latency_chart = LineChart()
    latency_chart.title = "Mean latency"
    latency_chart.y_axis.title = "milliseconds"
    latency_chart.add_data(
        Reference(sheet, min_col=5, min_row=start, max_row=last), titles_from_data=True
    )
    latency_chart.set_categories(
        Reference(sheet, min_col=1, max_col=2, min_row=start + 1, max_row=last)
    )
    latency_chart.height = 9
    latency_chart.width = 20
    sheet.add_chart(latency_chart, "J43")

    workbook.save(out)


def main():
    rows = read_rows(SOURCE)
    nodes = read_context(SOURCE)
    charts = HERE / "benchmark-results.png"
    terminal = HERE / "benchmark-console.png"
    workbook = HERE / "benchmark-results.xlsx"

    make_charts(rows, nodes, charts)
    make_terminal_screenshot(SOURCE, terminal)
    make_workbook(rows, nodes, workbook)

    for path in (charts, terminal, workbook):
        print(f"wrote {path}")


if __name__ == "__main__":
    main()
